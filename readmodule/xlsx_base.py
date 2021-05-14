#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@file        :xlsx_base.py
@Description:       :spcific class,
For xlsx files recommended use openpyxl to read and write.
This script use xlrd==1.2.0 which can read both xlsx and xls.
@Date     :2021/01/16 10:55:18
@Author      :hotwa
@version      :1.0
'''

import xlwt
import base64
from xlrd import open_workbook
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader # picture dependency
from openpyxl.utils import get_column_letter, column_index_from_string # number transform letter
from PIL import Image
from io import BytesIO
try:
    from xlsx_base_a import *
except ModuleNotFoundError as error:
    if __package__ is None:
        import sys
        sys.path.append('../../../')
    from readmodule.xlsx_base_a import *



class xlsx_base(casheet):
    __slots__ = [
        '__filename', 'label_model', 'xlsx_sheets_info', 'workbook',
        'label_sheet_mode', 'openpyxl_wb','all_sheets'
    ]

    def __init__(
        self,
        filename: dict(type=str, help="XLSX file name or path"),
        label_model: dict(type=str,help="Set default label model:row/col") = "row"):
        super().__init__(filename=filename) # 向上查找对应变量，至报错
        self.workbook = open_workbook(filename)
        self.openpyxl_wb = load_workbook(filename)
        self.all_sheets = self.openpyxl_wb.sheetnames
        self.__current_sheet_num = 0
        self.label_model = label_model
        self.xlsx_sheets_info = [[
            k, self.workbook.sheet_names()[k], v.nrows, v.ncols
        ] for k, v in enumerate(self.workbook.sheets())] # [[0, 'Sheet2', 105, 32], [序列, '表名', 行数, 列数]] 实例化读取xlsx所有信息

    
    @property
    def current_sheet_num(self): # current sheet num
        return self.__current_sheet_num

    @current_sheet_num.setter
    def current_sheet_num(self, value):
        if not isinstance(value, int):
            raise ValueError('current_sheet_num must be an integer!')
        self.__current_sheet_num = int(value)
        self.current_sheet_name = self.workbook.sheet_names()[
            self.__current_sheet_num] # 用户设置了表的序列，则同时将当前表名称属性设置

    
    @property
    def current_sheet_name(self): # 当前表名称属性,只有用户主动设置了当前表的编号时此属性才生效，默认使用第一个表的时候不生效
        return self.__current_sheet_name

    @current_sheet_name.setter
    def current_sheet_name(self, value):
        if not isinstance(value, str):
            raise ValueError('current_sheet_name must be a string!')
        self.__current_sheet_name = value

    def read_xlsx_info(self) -> list:
        """read_xlsx_info [read xlsx file sheets infos]
        method
        [print some details infos]
        
        :return: [struct of this current xlsx]
        :rtype: [list]
        """
        workbook = self.workbook
        print('current xlsx file：', self.filename)
        print('workbook number:', workbook.nsheets)
        [
            print(f'workbook name:{workbook.sheet_names()[k]};workbook rows:{worksheet.nrows};workbook columns:{worksheet.ncols}') for k, worksheet in enumerate(workbook.sheets())
        ]
        return self.xlsx_sheets_info

    def current_sheet_info(self):
        """current_sheet_info print some infos about current sheet
        
        [extended_summary]
        
        """
        workbook = self.workbook
        print(
            f'current sheet info:{workbook.sheet_names()[self.__current_sheet_num]}\nrow number：{ self.current_sheet_nrows}\ncolumns numbers：{self.current_sheet_ncols}'
        )

    def read_sheet_lable(
        self,
        sheet_num: dict(type=int,
                        help="A sequence of tables in an XLSX file") = None,
        **kwargs: dict(
            type=dict,
            help=
            "Include sheet_name and sheet_mode params,\n sheet_name:with sheet_name to read.\n sheet_mode:read xlsx sheet mode, choose first row or first column. col/row"
        )) -> list:
        """
        @description : 根据用户输入表名或者表的序列来读取表中第一列中值作为key值，并传入列表。
        ---------
        @param sheet_num: number, xlsx sheet number
        @param sheet_name: string, xlsx sheet name
        @param sheet_mode: string, read xlsx sheet mode, choose first row or first column. col/row as label in database
        -------
        @Returns list: The value of first row or first column
        -------
        [use xlrd libary]
        """
        _sheet_name = kwargs['sheet_name'] if ('sheet_name' in kwargs) else None # 获取表名
        if sheet_num == None: sheet_num = self.__current_sheet_num # 获取表序列，如果获取不到sheet_num，将当前设定的表锁定的序列传入，默认为0，第一个sheet
        label_model = kwargs['sheet_mode'] if ('sheet_mode' in kwargs) else 'row' # 获取读取表为第一行还是第一列读取模式,默认第一行读取
        self.label_model = label_model # 如果此处用户修改了默认以第一行作为标签则修改默认的读取方式
        # 优先以表名读取，否则默认读取设置的当前表顺序读取
        __xlsx_sheet_obj = self.workbook.sheet_by_name(_sheet_name) if (
            _sheet_name) else self.workbook.sheet_by_index(sheet_num)
        __sheet_rows = __xlsx_sheet_obj.nrows
        __sheet_cols = __xlsx_sheet_obj.ncols
        # 先从xlrd读取对象里面获取名称为sheet_name的表或者编号为sheet_num的表（default：0）
        # 读取此表的清除空值第一行（列）所有的值，传入first_row
        if label_model == 'col':
            first_row = [
                i for i in __xlsx_sheet_obj.col_values(colx=0) if i != ''
            ]
        else:
            first_row = [
                i for i in __xlsx_sheet_obj.row_values(rowx=0) if i != ''
            ]
        return first_row

    def read_sheet_line(self,
                        num: int,
                        mode='row') -> list:
        """read_sheet_line [read a row/columns]
        
        [read a sheet, the suquence of num lines(row,column),return a list]
        [use xlrd libary]
        :param num: [the read of col/row numbers]
        :type num: [int]
        :param mode: [col/row], defaults to 'row'
        :type mode: [string], optional
        :return: [list of the line]
        :rtype: [list]
        """
        workbook_obj = self.workbook
        workbook_sheet_obj = workbook_obj.sheet_by_index(
            self.__current_sheet_num)
        _line_list = workbook_sheet_obj.col_values(
            colx=num
        ) if self.label_model == 'col' else workbook_sheet_obj.row_values(
            rowx=num)
        return _line_list
    
    def get_cell_pic(self,position,new_name=None,base64c=None)->bytes:
        """get_cell_pic get_cell_pic get cell position picture, rename and save it, if picture not exists return None
        
        [use openpyxl libary]
        
        :param position: [picture cell]
        :type position: [string]
        :param new_name: [save filename]
        :type new_name: [string]
        """
        wb = load_workbook(self.filename)
        ws = wb.active # 调用得到正在运行的工作表
        ws.title # 调用当前运行的工作的名称
        sheet = wb[self.all_sheets[self.__current_sheet_num]] # read first sheet
        image_loader = SheetImageLoader(sheet)
        if image_loader.image_in(position):
            image = image_loader.get(position)
            if new_name:
                # 储存文件
                image.save(new_name)
            else:
                # 返回二进制内容
                bytesIO = BytesIO()
                image.save(bytesIO, format='PNG')
                if base64c:
                    return base64.b64encode(bytesIO.getvalue()).decode() # default to string
                    # return '1' # for test
                else:
                    return bytesIO.getvalue()
        else:
            return None

    def read_cell(self,pos:dict(type=str,help='excell position like E2')):
        """read_cell read_cell openpyxl methods get cell values
        
        [picture cell return bytes(format:PNG)]
        
        :param pos: [cell description], defaults to str,help='excell position like E2')
        :type pos: [string], optional
        """
        sheet = self.openpyxl_wb[self.all_sheets[self.__current_sheet_num]]
        image_loader = SheetImageLoader(sheet)
        if image_loader.image_in(pos): # picture cell return bytes
            bytes_value = self.get_cell_pic(pos)
            return bytes_value
        else:
            return sheet[pos].value

    def read_row(self,row_num:int,base64c=None,read_cell_picture: bool=False,pic_column:list=[]):
        """read_row 相比read_sheet_line方法，空cell返回None，默认数字类型自动转化为int
        
        [extended_summary]
        
        :param row_num: [行数]
        :type row_num: [int]
        :param read_cell_picture: [自动尝试读取单元格图片], defaults to False
        :type read_cell_picture: [bool], optional
        :param pic_column: [read columns of string list], defaults to []
        :type pic_column: [list], optional
        :return: [返回改行读取的列表]
        :rtype: [list]
        """
        sheet = self.openpyxl_wb[self.all_sheets[self.__current_sheet_num]] # open sheet
        row = sheet[row_num]
        row_num_sum = len(row)
        _row_list = []
        for i in range(row_num_sum):
            _row_list.append(row[i].value)
        _ll = [False if i == None else 1 for i in _row_list]
        if sum(_ll) == 0:
            return None # 此行为空行
        else:
            if read_cell_picture:
                if pic_column == []:
                    row_list = self.__cell_pic_try(row_num=row_num,current_row_list=_row_list,base64c=base64c)
                else:
                    row_list = self.__cell_pic_try(row_num=row_num,current_row_list=_row_list,pic_column=pic_column,base64c=base64c)
                return row_list
            return _row_list

    def __cell_pic_try(self,row_num,current_row_list,base64c=None,pic_column=[])-> list:
        """__cell_pic_try 尝试对改行的None进行图片提取返回二进制
        
        [extended_summary]
        
        :param row_num: [行号]
        :type row_num: [int]
        """
        # print(f'当前行读取内容{current_row_list},{row_num}')
        if None in current_row_list:
            while (None in current_row_list):
                _pos = current_row_list.index(None)
                letter = get_column_letter(_pos+1)
                pos=letter+str(row_num)
                # print(f'尝试定位{pos}位置的图片')
                if pic_column == []:
                    res = self.get_cell_pic(position=pos,base64c = base64c)
                    if res:
                        # print(f'定位图片成功{pos}')
                        current_row_list[_pos] = res
                    else:
                        current_row_list[_pos] = False
                elif letter in pic_column:
                    res = self.get_cell_pic(position=pos,base64c = base64c)
                    if res:
                        # print(f'定位图片成功{pos}')
                        current_row_list[_pos] = res
                    else:
                        current_row_list[_pos] = False
                else:
                    # print(f'跳过{pos}位置的图片')
                    current_row_list[_pos] = False
            return current_row_list
        else:
            return None

if __name__ == '__main__':
    a=xlsx_base(filename='./v3_20210330remove_empty.xlsx')
